"""Automated business reports: PDF (matplotlib PdfPages) and XLSX (openpyxl).

Matplotlib over WeasyPrint deliberately: pure-Python wheels, no system font/
pango dependencies, renders identically in Docker (docs/plan/phase-5 risk table).
"""

import io
from datetime import date

import matplotlib
import pandas as pd
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
from matplotlib.backends.backend_pdf import PdfPages  # noqa: E402


async def _report_data(db: AsyncSession, start: date, end: date) -> dict:
    kpis = pd.DataFrame(
        (
            await db.execute(
                text(
                    "SELECT metric, snapshot_date, value FROM kpi_snapshots "
                    "WHERE dimensions = '{}'::jsonb AND snapshot_date BETWEEN :s AND :e "
                    "ORDER BY snapshot_date"
                ),
                {"s": start, "e": end},
            )
        ).all(),
        columns=["metric", "date", "value"],
    )
    top_products = (
        await db.execute(
            text(
                "SELECT p.name, SUM(st.total_amount) AS revenue, SUM(st.quantity) AS qty "
                "FROM sales_transactions st JOIN products p ON p.id = st.product_id "
                "WHERE st.txn_date BETWEEN :s AND :e "
                "GROUP BY p.name ORDER BY revenue DESC LIMIT 8"
            ),
            {"s": start, "e": end},
        )
    ).all()
    insights = (
        await db.execute(
            text(
                "SELECT title, body, severity FROM insights "
                "WHERE generated_at::date BETWEEN :s AND :e "
                "ORDER BY generated_at DESC LIMIT 8"
            ),
            {"s": start, "e": end},
        )
    ).all()
    anomalies = (
        await db.execute(
            text(
                "SELECT metric, severity, status, context->>'date' AS day, "
                "observed_value, expected_value FROM anomalies "
                "WHERE (context->>'date')::date BETWEEN :s AND :e ORDER BY day"
            ),
            {"s": start, "e": end},
        )
    ).all()
    return {
        "kpis": kpis,
        "top_products": top_products,
        "insights": insights,
        "anomalies": anomalies,
    }


def _kpi_totals(kpis: pd.DataFrame) -> dict[str, float]:
    if kpis.empty:
        return {}
    sums = kpis.groupby("metric")["value"].sum()
    return {
        m: float(sums.get(m, 0)) for m in ("revenue", "orders", "expense_total", "gross_margin")
    }


async def build_pdf(db: AsyncSession, start: date, end: date, title: str) -> bytes:
    data = await _report_data(db, start, end)
    totals = _kpi_totals(data["kpis"])
    buf = io.BytesIO()

    with PdfPages(buf) as pdf:
        # Page 1 — headline
        fig = plt.figure(figsize=(8.27, 11.69))  # A4
        fig.suptitle(title, fontsize=16, fontweight="bold", y=0.97)
        fig.text(0.5, 0.935, f"Period: {start:%d %b %Y} — {end:%d %b %Y}", ha="center", fontsize=10)

        rows = [
            ["Revenue", f"NPR {totals.get('revenue', 0):,.0f}"],
            ["Orders", f"{totals.get('orders', 0):,.0f}"],
            ["Gross margin", f"NPR {totals.get('gross_margin', 0):,.0f}"],
            ["Expenses", f"NPR {totals.get('expense_total', 0):,.0f}"],
            [
                "Net (margin − expenses)",
                f"NPR {totals.get('gross_margin', 0) - totals.get('expense_total', 0):,.0f}",
            ],
        ]
        ax = fig.add_axes((0.1, 0.62, 0.8, 0.25))
        ax.axis("off")
        table = ax.table(cellText=rows, colWidths=[0.5, 0.5], loc="center", cellLoc="left")
        table.scale(1, 1.8)
        table.auto_set_font_size(False)
        table.set_fontsize(11)

        rev = data["kpis"][data["kpis"]["metric"] == "revenue"]
        if not rev.empty:
            ax2 = fig.add_axes((0.1, 0.28, 0.8, 0.28))
            ax2.plot(rev["date"], rev["value"], color="#4f46e5", linewidth=1.5)
            ax2.set_title("Daily revenue", fontsize=11)
            ax2.tick_params(labelsize=8)
            for spine in ("top", "right"):
                ax2.spines[spine].set_visible(False)

        if data["top_products"]:
            ax3 = fig.add_axes((0.1, 0.04, 0.8, 0.18))
            names = [r.name[:24] for r in data["top_products"]][::-1]
            values = [float(r.revenue) for r in data["top_products"]][::-1]
            ax3.barh(names, values, color="#10b981")
            ax3.set_title("Top products by revenue", fontsize=11)
            ax3.tick_params(labelsize=8)
        pdf.savefig(fig)
        plt.close(fig)

        # Page 2 — insights & anomalies
        fig = plt.figure(figsize=(8.27, 11.69))
        fig.suptitle("Insights & anomalies", fontsize=14, fontweight="bold", y=0.97)
        y = 0.90
        fig.text(0.08, y, "Automated insights", fontsize=12, fontweight="bold")
        y -= 0.03
        if not data["insights"]:
            fig.text(0.08, y, "No insights generated in this period.", fontsize=9)
            y -= 0.03
        for title_, body, severity in data["insights"]:
            fig.text(0.08, y, f"• [{severity}] {title_}", fontsize=9.5, fontweight="bold")
            y -= 0.022
            import textwrap

            for line in textwrap.wrap(body, 100)[:3]:
                fig.text(0.10, y, line, fontsize=8.5)
                y -= 0.018
            y -= 0.012
            if y < 0.30:
                break
        fig.text(0.08, 0.26, "Anomalies in period", fontsize=12, fontweight="bold")
        y = 0.23
        if not data["anomalies"]:
            fig.text(0.08, y, "No anomalies detected in this period.", fontsize=9)
        for metric, severity, status_, day, observed, expected in data["anomalies"][:8]:
            fig.text(
                0.08,
                y,
                f"• {day}  {metric}: observed NPR {float(observed):,.0f} vs expected "
                f"NPR {float(expected or 0):,.0f}  [{severity}, {status_}]",
                fontsize=8.5,
            )
            y -= 0.02
        pdf.savefig(fig)
        plt.close(fig)

    return buf.getvalue()


async def build_xlsx(db: AsyncSession, start: date, end: date) -> bytes:
    from openpyxl import Workbook

    data = await _report_data(db, start, end)
    wb = Workbook()

    ws = wb.active
    ws.title = "KPIs"
    ws.append(["Metric", "Total"])
    for metric, total in _kpi_totals(data["kpis"]).items():
        ws.append([metric, total])

    ws2 = wb.create_sheet("Top products")
    ws2.append(["Product", "Revenue (NPR)", "Quantity"])
    for r in data["top_products"]:
        ws2.append([r.name, float(r.revenue), int(r.qty)])

    ws3 = wb.create_sheet("Daily revenue")
    ws3.append(["Date", "Revenue"])
    rev = data["kpis"][data["kpis"]["metric"] == "revenue"]
    for _, row in rev.iterrows():
        ws3.append([row["date"].isoformat(), float(row["value"])])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
